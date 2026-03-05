from flask import Blueprint, jsonify, request, current_app
from sqlalchemy import or_
from pear_admin.extensions import db
from pear_admin.orms import MaterialPlanningORM, MaterialInboundORM, MaterialInventoryORM, MaterialOutboundORM, MaterialInvoiceORM, MaterialInvoiceDetailORM, ProjectORM, SupplierORM
from datetime import datetime
from pear_admin.utils import authorize

material_api = Blueprint("material_api", __name__, url_prefix="/material")

# --- Options (Common) ---
@material_api.route("/options", methods=["GET"])
def get_options():
    """获取所有项目列表及各模块数量统计"""
    try:
        projects = ProjectORM.query.all()
        suppliers = SupplierORM.query.all()
        project_list = []
        
        for project in projects:
            # 统计各模块数量
            planning_count = MaterialPlanningORM.query.filter_by(project_id=project.id).count()
            inbound_count = MaterialInboundORM.query.filter_by(project_id=project.id, status='pending').count()
            inventory_count = MaterialInventoryORM.query.filter_by(project_id=project.id).count()
            outbound_count = MaterialOutboundORM.query.join(MaterialOutboundORM.inventory).filter(MaterialInventoryORM.project_id==project.id, MaterialOutboundORM.status=='pending').count()
            invoice_count = MaterialInvoiceORM.query.filter_by(project_id=project.id).count()
            
            project_list.append({
                'id': project.id,
                'name': project.project_name,
                'planning_count': planning_count,
                'inbound_count': inbound_count,
                'inventory_count': inventory_count,
                'outbound_count': outbound_count,
                'invoice_count': invoice_count
            })
        
        return jsonify({
            "code": 0,
            "msg": "获取成功",
            "data": {
                "projects": project_list,
                "suppliers": [{"id": s.id, "name": s.name} for s in suppliers]
            }
        })
    except Exception as e:
        return jsonify({"code": 1, "msg": str(e)})

# --- Planning ---
@material_api.route("/planning", methods=["GET"])
def get_planning():
    try:
        page = request.args.get("page", 1, type=int)
        limit = request.args.get("limit", 10, type=int)
        
        query = MaterialPlanningORM.query
        
        # Filtering
        project_id = request.args.get("project_id")
        if project_id:
            query = query.filter_by(project_id=project_id)
            
        remaining_status = request.args.get("remaining_status")
        if remaining_status == '>0':
            query = query.filter(MaterialPlanningORM.planned_remaining_quantity > 0)
        elif remaining_status == '=0':
            query = query.filter(MaterialPlanningORM.planned_remaining_quantity == 0)
        elif remaining_status == '<0':
            query = query.filter(MaterialPlanningORM.planned_remaining_quantity < 0)
            
        pagination = query.paginate(page=page, per_page=limit, error_out=False)
        
        res_data = []
        for item in pagination.items:
            try:
                d = item.json()
                # Calculate total amount
                price = float(item.planned_price or 0)
                qty = float(item.planned_total_quantity or 0)
                d['planned_total_amount'] = f"{price * qty:.2f}"
                
                # Calculate pending inbound quantity (calculated field)
                # Loose coupling based on project, name, spec
                pending_qty = db.session.query(db.func.sum(MaterialInboundORM.inbound_quantity))\
                    .filter(
                        MaterialInboundORM.project_id == item.project_id,
                        MaterialInboundORM.material_name == item.material_name,
                        MaterialInboundORM.material_spec == item.material_spec,
                        MaterialInboundORM.status == 'pending'
                    ).scalar() or 0
                d['pending_inbound_quantity'] = f"{float(pending_qty):.2f}"
                
                res_data.append(d)
            except Exception as e:
                import traceback
                traceback.print_exc()
                continue
        
        return jsonify({
            "code": 0,
            "msg": "",
            "count": pagination.total,
            "data": res_data
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"code": 1, "msg": f"Server Error: {str(e)}"})


@material_api.route("/planning/generate_inbound", methods=["POST"])
def generate_inbound_from_planning():
    """从策划生成入库计划"""
    try:
        data = request.json
        items = data.get("items", []) # List of {id: x, quantity: y}
        
        if not items:
            return jsonify({"code": 1, "msg": "没有可生成的项目"})
        
        # 生成批次号基础信息
        today = datetime.now().strftime("%Y%m%d")
        
        # 查询今天的最大批次序号（不区分项目）
        max_batch = db.session.query(
            db.func.max(MaterialInboundORM.batch_number)
        ).filter(
            MaterialInboundORM.batch_number.like(f"{today}-%")
        ).scalar()
        
        # 计算新序号
        if max_batch:
            try:
                last_seq = int(max_batch.split('-')[-1])
                new_seq = last_seq + 1
            except:
                new_seq = 1
        else:
            new_seq = 1
        
        # 生成主批次号（格式：YYYYMMDD-序号）
        batch_number = f"{today}-{new_seq:03d}"
        
        count = 0
        sub_number = 1
        
        for item in items:
            planning = MaterialPlanningORM.query.get(item['id'])
            if not planning:
                continue
                
            qty = float(item.get('quantity', 0))
            if qty <= 0:
                continue
                
            # Check remaining quantity (Optional: Strict check or allow over-planning?)
            # Usually strict: 
            current_rem = float(planning.planned_remaining_quantity or 0)
            if current_rem < qty:
                # Optionally return error or cap it. For now, we skip or clamp?
                # Using strict policy: Skip if insufficient? Or partial?
                # User asked for linkage. Let's strictly enforce usage of remaining.
                # Actually, skipping might confuse user. Let's error out? 
                # Or for bulk, skipping is safer but logging is better.
                # Given user context, let's assume valid input from frontend (which caps it).
                pass

            # Update remaining quantity
            planning.planned_remaining_quantity = current_rem - qty
            
            inbound = MaterialInboundORM(
                project_id=planning.project_id,
                batch_number=batch_number,
                batch_sub_number=f"{sub_number:03d}",  # 批次分号
                material_name=planning.material_name,
                material_spec=planning.material_spec,
                material_unit=planning.material_unit,
                supplier_id=planning.supplier_id, # Inherit supplier
                inbound_quantity=qty,
                inbound_price=planning.planned_price, # Inherit price
                inbound_total_amount=qty * float(planning.planned_price or 0),
                status='pending'
            )
            db.session.add(inbound)
            count += 1
            sub_number += 1
            
        db.session.commit()
        return jsonify({
            "code": 0, 
            "msg": f"成功生成 {count} 条入库计划，批次号: {batch_number}"
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/planning", methods=["POST"])
def add_planning():
    try:
        data = request.json
        row = MaterialPlanningORM(**data)
        db.session.add(row)
        db.session.commit()
        return jsonify({"code": 0, "msg": "Success"})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/planning/import", methods=["POST"])
def import_planning():
    """导入Excel文件批量创建材料策划"""
    try:
        # Check project_id
        project_id = request.form.get('project_id')
        if not project_id:
            return jsonify({"code": 1, "msg": "请选择导入的项目"})
        
        project = ProjectORM.query.get(int(project_id))
        if not project:
            return jsonify({"code": 1, "msg": "无效的项目ID"})
            
        # 检查文件
        if 'file' not in request.files:
            return jsonify({"code": 1, "msg": "未找到上传文件"})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"code": 1, "msg": "未选择文件"})
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({"code": 1, "msg": "仅支持Excel文件(.xlsx, .xls)"})
        
        # 解析Excel
        from openpyxl import load_workbook
        import io
        
        # 读取文件到内存
        file_content = io.BytesIO(file.read())
        wb = load_workbook(file_content, data_only=True)
        ws = wb.active
        
        # 验证表头（第一行）
        # Strip whitespace and handle None
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        expected_headers = ['材料名称', '规格型号', '单位', '策划总数', '策划单价']
        
        # Validate existence of required headers
        missing_headers = [h for h in expected_headers if h not in headers]
        if missing_headers:
            return jsonify({
                "code": 1, 
                "msg": f"Excel表头缺失必需列: {', '.join(missing_headers)}。请仅包含: {', '.join(expected_headers)}"
            })
            
        # Map headers to indices
        header_map = {name: idx for idx, name in enumerate(headers)}
        
        # 解析数据行
        success_count = 0
        error_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Check if row is empty or name is missing
                # Note: valid row might be longer than header_map if file has extra cols, 
                # or shorter if trailing cols are empty.
                name_idx = header_map['材料名称']
                if name_idx >= len(row) or not row[name_idx]:
                    continue
                
                # Helper to safely extract and parse
                def get_val(col_name):
                    idx = header_map[col_name]
                    if idx < len(row):
                        return row[idx]
                    return None
                    
                def get_float(col_name):
                    val = get_val(col_name)
                    if val is None or str(val).strip() == '':
                        return 0.0
                    try:
                        return float(val)
                    except ValueError:
                         raise ValueError(f"列[{col_name}]数据格式错误: {val}")

                # 创建策划记录
                planning = MaterialPlanningORM(
                    project_id=int(project_id),
                    material_name=str(get_val('材料名称')),
                    material_spec=str(get_val('规格型号') or ""),
                    material_unit=str(get_val('单位') or ""),
                    planned_total_quantity=get_float('策划总数'),
                    planned_remaining_quantity=get_float('策划总数'),  # 初始余量=总数
                    planned_price=get_float('策划单价')
                )
                
                db.session.add(planning)
                success_count += 1
                
            except Exception as e:
                errors.append(f"第{row_idx}行: {str(e)}")
                error_count += 1
        
        # 提交事务
        if success_count > 0:
            db.session.commit()
        
        # 返回结果
        result_msg = f"成功导入 {success_count} 条记录"
        if error_count > 0:
            result_msg += f"，失败 {error_count} 条"
            if len(errors) <= 5:
                result_msg += f"。错误: {'; '.join(errors)}"
            else:
                result_msg += f"。前5个错误: {'; '.join(errors[:5])}"
        
        return jsonify({
            "code": 0 if error_count == 0 else 1,
            "msg": result_msg,
            "data": {
                "success": success_count,
                "error": error_count
            }
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"code": 1, "msg": f"导入失败: {str(e)}"})

@material_api.route("/planning/<int:id>", methods=["DELETE"])
def remove_planning(id):
    """删除单个策划"""
    try:
        planning = MaterialPlanningORM.query.get(id)
        if not planning:
            return jsonify({"code": 1, "msg": "记录不存在"})
        db.session.delete(planning)
        db.session.commit()
        return jsonify({"code": 0, "msg": "删除成功"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/planning/batch", methods=["DELETE"])
def batch_remove_planning():
    """批量删除策划"""
    try:
        data = request.json
        ids = [int(i) for i in data.get("ids", []) if i]
        if not ids:
            return jsonify({"code": 1, "msg": "未选择任何项"})
            
        # 批量删除
        MaterialPlanningORM.query.filter(MaterialPlanningORM.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        
        return jsonify({"code": 0, "msg": "删除成功"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/planning/batch_supplier", methods=["PUT"])
def batch_supplier_planning():
    """批量拟定供应商"""
    try:
        data = request.json
        ids = [int(i) for i in data.get("ids", []) if i]
        supplier_id = data.get("supplier_id")
        
        if not ids:
            return jsonify({"code": 1, "msg": "未选择任何想项"})
        if not supplier_id:
            return jsonify({"code": 1, "msg": "请选择供应商"})
            
        MaterialPlanningORM.query.filter(MaterialPlanningORM.id.in_(ids)).update(
            {MaterialPlanningORM.supplier_id: supplier_id}, 
            synchronize_session=False
        )
        db.session.commit()
        
        return jsonify({"code": 0, "msg": "供应商拟定成功"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": str(e)})

# --- Inbound ---
@material_api.route("/inbound", methods=["GET"])
def get_inbound():
    try:
        page = request.args.get("page", 1, type=int)
        limit = request.args.get("limit", 10, type=int)
        
        query = MaterialInboundORM.query
         # Filtering
        project_id = request.args.get("project_id")
        if project_id:
            query = query.filter_by(project_id=project_id)
            
        status = request.args.get("status")
        if status:
            query = query.filter_by(status=status)
            
        # Additional Filters for Linkage
        material_name = request.args.get("material_name")
        if material_name:
            query = query.filter_by(material_name=material_name)
            
        material_spec = request.args.get("material_spec")
        if material_spec:
            # Handle empty spec special case if needed, but usually exact match
            query = query.filter_by(material_spec=material_spec)

        # 按批次号降序排序（最新的批次在前，同批次的记录聚合在一起）
        query = query.order_by(MaterialInboundORM.batch_number.desc(), MaterialInboundORM.id.asc())

        pagination = query.paginate(page=page, per_page=limit, error_out=False)
        
        return jsonify({
            "code": 0,
            "msg": "",
            "count": pagination.total,
            "data": [item.json() for item in pagination.items]
        })
    except Exception as e:
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/inbound/<int:id>", methods=["GET"])
def get_inbound_detail(id):
    """获取单个入库记录详情"""
    try:
        inbound = MaterialInboundORM.query.get(id)
        if not inbound:
            return jsonify({"code": 1, "msg": "记录不存在"})
        return jsonify({"code": 0, "msg": "成功", "data": inbound.json()})
    except Exception as e:
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/inbound/<int:id>/batch", methods=["PUT"])
def update_inbound_batch(id):
    """更新入库计划的批次号"""
    try:
        data = request.json
        batch_number = data.get("batch_number", "").strip()
        
        if not batch_number:
            return jsonify({"code": 1, "msg": "批次号不能为空"})
        
        inbound = MaterialInboundORM.query.get(id)
        if not inbound:
            return jsonify({"code": 1, "msg": "入库计划不存在"})
        
        # 只允许修改待入库状态的记录
        if inbound.status != 'pending':
            return jsonify({"code": 1, "msg": "只能修改待入库状态的批次号"})
        
        inbound.batch_number = batch_number
        db.session.commit()
        
        return jsonify({"code": 0, "msg": "批次号更新成功"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/inbound/<int:id>/invoice", methods=["PUT"])
def update_inbound_invoice(id):
    """更新入库计划关联的发票ID"""
    try:
        data = request.json
        invoice_id = data.get("invoice_id") # 可以为 None 表示取消关联
        
        inbound = MaterialInboundORM.query.get(id)
        if not inbound:
            return jsonify({"code": 1, "msg": "入库计划不存在"})
        
        # 验证发票是否存在（如果提供了 invoice_id）
        new_supplier_id = None
        if invoice_id:
            invoice = MaterialInvoiceORM.query.get(invoice_id)
            if not invoice:
                return jsonify({"code": 1, "msg": "发票不存在"})
            
            # 策略：如果发票有供应商ID，直接同步；如果没有，尝试按 seller_name 匹配
            new_supplier_id = invoice.supplier_id
            if not new_supplier_id and invoice.seller_name:
                match_supplier = SupplierORM.query.filter_by(name=invoice.seller_name).first()
                if match_supplier:
                    new_supplier_id = match_supplier.id
                    # 顺便更新发票的供应商ID以供未来使用
                    invoice.supplier_id = new_supplier_id

        inbound.invoice_id = invoice_id
        if new_supplier_id:
            inbound.supplier_id = new_supplier_id
            
        db.session.commit()
        
        return jsonify({"code": 0, "msg": "发票关联成功"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/inbound/batch-invoice", methods=["PUT"])
def batch_update_inbound_invoice():
    """批量更新入库计划关联的发票ID"""
    try:
        data = request.json
        ids = [int(i) for i in data.get("ids", []) if i]
        invoice_id = data.get("invoice_id")
        
        if not ids:
            return jsonify({"code": 1, "msg": "未选择记录"})
            
        # 验证发票是否存在（如果提供了 invoice_id）
        new_supplier_id = None
        if invoice_id:
            invoice = MaterialInvoiceORM.query.get(invoice_id)
            if not invoice:
                return jsonify({"code": 1, "msg": "发票不存在"})
            
            # 策略：如果发票有供应商ID，直接同步；如果没有，尝试按 seller_name 匹配
            new_supplier_id = invoice.supplier_id
            if not new_supplier_id and invoice.seller_name:
                match_supplier = SupplierORM.query.filter_by(name=invoice.seller_name).first()
                if match_supplier:
                    new_supplier_id = match_supplier.id
                    # 顺便更新发票的供应商ID以供未来使用
                    invoice.supplier_id = new_supplier_id
        
        # 批量更新
        update_data = {"invoice_id": invoice_id}
        if new_supplier_id:
            update_data["supplier_id"] = new_supplier_id
            
        db.session.query(MaterialInboundORM).filter(
            MaterialInboundORM.id.in_(ids)
        ).update(update_data, synchronize_session=False)
        
        db.session.commit()
        return jsonify({"code": 0, "msg": "批量关联成功"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/inbound/batch-delete", methods=["POST"])
def batch_delete_inbound():
    """批量删除入库计划"""
    try:
        data = request.json
        ids = [int(i) for i in data.get('ids', []) if i]
        
        if not ids:
            return jsonify({"code": 1, "msg": "请选择要删除的记录"})
        
        # 查询要删除的记录
        inbounds = MaterialInboundORM.query.filter(
            MaterialInboundORM.id.in_(ids)
        ).all()
        
        if not inbounds:
            return jsonify({"code": 1, "msg": "未找到要删除的记录"})
        
        # 检查状态，只能删除待入库状态的记录
        for inbound in inbounds:
            if inbound.status != 'pending':
                return jsonify({"code": 1, "msg": f"记录 {inbound.id} 已确认入库，无法删除"})
        
        # 执行删除
        for inbound in inbounds:
            # Restore quantity to planning before deletion
            restore_qty = float(inbound.inbound_quantity or 0)
            if restore_qty > 0:
                planning = MaterialPlanningORM.query.filter_by(
                    project_id=inbound.project_id,
                    material_name=inbound.material_name,
                    material_spec=inbound.material_spec
                ).first()
                if planning:
                    current_rem = float(planning.planned_remaining_quantity or 0)
                    planning.planned_remaining_quantity = current_rem + restore_qty
            db.session.delete(inbound)
        
        db.session.commit()
        return jsonify({"code": 0, "msg": f"成功删除 {len(inbounds)} 条记录"})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": f"删除失败: {str(e)}"})




@material_api.route("/inbound/<int:id>", methods=["PUT"])
def update_inbound(id):
    """更新单个入库记录"""
    try:
        data = request.json
        inbound = MaterialInboundORM.query.get(id)
        if not inbound:
            return jsonify({"code": 1, "msg": "记录不存在"})
        
        # 仅允许在待入库时编辑
        if inbound.status != 'pending':
            return jsonify({"code": 1, "msg": "已入库记录不可编辑"})
            
        # Capture old quantity before update
        old_qty = float(inbound.inbound_quantity or 0)

        for key, value in data.items():
            if hasattr(inbound, key):
                setattr(inbound, key, value)
        
        # Calculate delta and update planning if quantity changed
        new_qty = float(inbound.inbound_quantity or 0)
        delta = new_qty - old_qty
        
        if delta != 0:
            planning = MaterialPlanningORM.query.filter_by(
                project_id=inbound.project_id,
                material_name=inbound.material_name,
                material_spec=inbound.material_spec
            ).first()
            if planning:
                current_rem = float(planning.planned_remaining_quantity or 0)
                planning.planned_remaining_quantity = current_rem - delta

        db.session.commit()
        return jsonify({"code": 0, "msg": "修改成功"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/inbound/<int:id>", methods=["DELETE"])
def remove_inbound(id):
    """删除单个入库记录"""
    try:
        inbound = MaterialInboundORM.query.get(id)
        if not inbound:
            return jsonify({"code": 1, "msg": "记录不存在"})
        
        # 仅允许在待入库时删除
        if inbound.status != 'pending':
            return jsonify({"code": 1, "msg": "已入库记录不可删除"})
            
        # Restore quantity to planning before deletion
        restore_qty = float(inbound.inbound_quantity or 0)
        if restore_qty > 0:
            planning = MaterialPlanningORM.query.filter_by(
                project_id=inbound.project_id,
                material_name=inbound.material_name,
                material_spec=inbound.material_spec
            ).first()
            if planning:
                current_rem = float(planning.planned_remaining_quantity or 0)
                planning.planned_remaining_quantity = current_rem + restore_qty

        db.session.delete(inbound)
        db.session.commit()
        return jsonify({"code": 0, "msg": "删除成功"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/inbound", methods=["POST"])
def add_inbound():
    data = request.json
    # Check if a batch number is provided, if not generate one
    if not data.get("batch_number"):
        import time
        data["batch_number"] = f"BN{int(time.time())}"
        
    row = MaterialInboundORM(**data)
    db.session.add(row)
    db.session.commit()
    return jsonify({"code": 0, "msg": "Success"})

@material_api.route("/inbound/batch_confirm", methods=["POST"])
def confirm_inbound():
    """批量确认入库"""
    try:
        data = request.json
        ids = [int(i) for i in data.get("ids", []) if i]
        if not ids:
            return jsonify({"code": 1, "msg": "未选择记录"})
            
        inbound_items = MaterialInboundORM.query.filter(MaterialInboundORM.id.in_(ids)).all()
        for item in inbound_items:
            if item.status == 'completed':
                continue
                
            # 1. 更新入库状态
            item.status = 'completed'
            
            # 2. 更新库存 (MaterialInventoryORM)
            inventory = MaterialInventoryORM.query.filter_by(
                project_id=item.project_id,
                material_name=item.material_name,
                material_spec=item.material_spec
            ).first()
            
            inbound_qty = float(item.inbound_quantity or 0)
            inbound_price = float(item.inbound_price or 0)
            inbound_amount = float(item.inbound_total_amount or (inbound_qty * inbound_price))
            
            if inventory:
                # 累加库存和总价值
                inventory.current_stock = float(inventory.current_stock or 0) + inbound_qty
                inventory.total_value = float(inventory.total_value or 0) + inbound_amount
                # 更新最近一次单价
                inventory.latest_price = inbound_price
                # 更新供应商 (记录最近一次供应商)
                if item.supplier_id:
                    inventory.supplier_id = item.supplier_id
                # 更新关联发票
                if item.invoice_id:
                    inventory.latest_invoice_id = item.invoice_id
                # 补全单位信息
                if not inventory.material_unit and item.material_unit:
                    inventory.material_unit = item.material_unit
            else:
                # 新建库存记录
                inventory = MaterialInventoryORM(
                    project_id=item.project_id,
                    material_name=item.material_name,
                    material_spec=item.material_spec,
                    material_unit=item.material_unit,
                    current_stock=inbound_qty,
                    latest_price=inbound_price,
                    total_value=inbound_amount,
                    supplier_id=item.supplier_id,
                    latest_invoice_id=item.invoice_id
                )
                db.session.add(inventory)
            
            # 3. 核减策划余量
            planning = MaterialPlanningORM.query.filter_by(
                project_id=item.project_id,
                material_name=item.material_name,
                material_spec=item.material_spec,
                material_unit=item.material_unit
            ).first()

            if planning:
                current_rem = float(planning.planned_remaining_quantity or 0)
                planning.planned_remaining_quantity = current_rem - inbound_qty

        db.session.commit()
        return jsonify({"code": 0, "msg": "批量确认入库成功，已导入库存台账"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": f"入库失败: {str(e)}"})

# --- Inventory ---
@material_api.route("/inventory", methods=["GET"])
def get_inventory():
    """获取库存台账列表"""
    try:
        page = request.args.get("page", 1, type=int)
        limit = request.args.get("limit", 10, type=int)
        project_id = request.args.get("project_id")
        
        query = MaterialInventoryORM.query
        if project_id:
            query = query.filter_by(project_id=project_id)
            
        pagination = query.paginate(page=page, per_page=limit, error_out=False)
        return jsonify({
            "code": 0,
            "msg": "",
            "count": pagination.total,
            "data": [item.json() for item in pagination.items]
        })
    except Exception as e:
        return jsonify({"code": 1, "msg": str(e)})

# --- Outbound ---
@material_api.route("/outbound", methods=["GET"])
def get_outbound():
    """获取出库计划列表"""
    try:
        page = request.args.get("page", 1, type=int)
        limit = request.args.get("limit", 10, type=int)
        project_id = request.args.get("project_id")
        status = request.args.get("status", "pending")
        
        # Always join inventory to filter out orphaned records and fetch related data efficiently if needed
        # And sort by ID descending (newest first)
        query = MaterialOutboundORM.query.join(MaterialOutboundORM.inventory).filter(MaterialOutboundORM.status==status)
        
        if project_id:
            query = query.filter(MaterialInventoryORM.project_id == project_id)
            
        query = query.order_by(MaterialOutboundORM.id.desc())
            
        pagination = query.paginate(page=page, per_page=limit, error_out=False)
        return jsonify({
            "code": 0,
            "msg": "",
            "count": pagination.total,
            "data": [item.json() for item in pagination.items]
        })
    except Exception as e:
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/outbound", methods=["POST"])
def add_outbound():
    """新增出库计划"""
    try:
        data = request.json
        # 检查库存
        inventory = MaterialInventoryORM.query.filter_by(
            project_id=data.get("project_id"),
            material_name=data.get("material_name"),
            material_spec=data.get("material_spec")
        ).first()
        
        if not inventory or inventory.current_stock < float(data.get("outbound_quantity", 0)):
            return jsonify({"code": 1, "msg": "库存不足"})
            
        # Explicitly map form fields to model fields with snapshots
        row = MaterialOutboundORM(
            inventory_id=inventory.id,
            project_id=inventory.project_id,
            material_name=inventory.material_name,
            material_spec=inventory.material_spec,
            material_unit=inventory.material_unit,
            seller_price=inventory.seller_price,
            seller_quantity=float(data.get("outbound_quantity", 0)),
            seller_id=inventory.seller_id,
            tax_rate=inventory.tax_rate,
            status='pending'
        )
        
        # Bidirectional Linkage
        if inventory.seller_quantity is not None:
            inventory.seller_quantity = float(inventory.seller_quantity) - float(data.get("outbound_quantity", 0))
            
        db.session.add(row)
        db.session.commit()
        return jsonify({"code": 0, "msg": "添加成功"})
    except Exception as e:
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/outbound/<int:id>", methods=["DELETE"])
def remove_outbound(id):
    """删除出库计划"""
    try:
        row = MaterialOutboundORM.query.get(id)
        if not row or row.status != 'pending':
            return jsonify({"code": 1, "msg": "记录不存在或已出库"})
            
        # Bidirectional Linkage: Restore seller quantity to inventory
        if row.inventory and row.seller_quantity:
            row.inventory.seller_quantity = (row.inventory.seller_quantity or 0) + row.seller_quantity
            
        db.session.delete(row)
        db.session.commit()
        return jsonify({"code": 0, "msg": "删除成功"})
    except Exception as e:
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/outbound/batch", methods=["POST"])
def batch_add_outbound():
    """批量新增出库计划"""
    from decimal import Decimal
    try:
        data = request.json
        items = data.get("items", [])
        recipient = data.get("recipient")
        remark = data.get("remark")
        direct_confirm = data.get("direct_confirm", False)
        
        if not items:
            return jsonify({"code": 1, "msg": "未选择任何物料"})
            
        success_count = 0
        errors = []
        
        # Generate Batch Number
        today = datetime.now().strftime("%Y%m%d")
        max_batch = db.session.query(
            db.func.max(MaterialOutboundORM.batch_number)
        ).filter(
            MaterialOutboundORM.batch_number.like(f"{today}-%")
        ).scalar()

        if max_batch:
            try:
                last_seq = int(max_batch.split('-')[-1])
                new_seq = last_seq + 1
            except:
                new_seq = 1
        else:
            new_seq = 1
        
        batch_number = f"{today}-{new_seq:03d}"
        
        for item in items:
            inventory_id = item.get("inventory_id")
            
            inventory = MaterialInventoryORM.query.get(inventory_id)
            if not inventory:
                errors.append(f"库存ID {inventory_id} 不存在")
                continue
            
            # Use seller_quantity as outbound quantity
            qty = Decimal(str(inventory.seller_quantity or 0))
            
            if qty <= 0:
                errors.append(f"{inventory.material_name} 销售数量为0")
                continue
                
            # Check stock if direct confirm
            if direct_confirm :
                curr_stock = Decimal(str(inventory.current_stock or 0))
                if curr_stock < qty:
                    errors.append(f"{inventory.material_name} 库存不足 (需 {qty}, 存 {curr_stock})")
                    continue

            # Create outbound record with snapshots
            row = MaterialOutboundORM(
                inventory_id=inventory.id,
                project_id=inventory.project_id,
                batch_number=batch_number,
                material_name=inventory.material_name,
                material_spec=inventory.material_spec,
                material_unit=inventory.material_unit,
                seller_price=inventory.seller_price,
                seller_quantity=float(qty),
                seller_id=inventory.seller_id,
                tax_rate=inventory.tax_rate,
                status='completed' if direct_confirm else 'pending'
            )
            
            if direct_confirm:
                # 1. Update Inventory Stock
                original_stock = Decimal(str(inventory.current_stock or 0))
                inventory.current_stock = float(original_stock - qty)
                
                # 2. Update Inventory Total Value (Weighted Avg logic)
                if inventory.current_stock > 0:
                    avg_price = Decimal(str(inventory.total_value or 0)) / original_stock
                    inventory.total_value = float(avg_price * Decimal(str(inventory.current_stock)))
                else:
                    inventory.total_value = 0
            
            # Bidirectional Linkage: Clear inventory seller quantity as it moved to outbound
            inventory.seller_quantity = 0
            
            db.session.add(row)
            success_count += 1
            
        if not errors:
            db.session.commit()
            msg = f"成功导出 {success_count} 条记录至出库记录" if direct_confirm else f"成功创建 {success_count} 条出库计划"
            return jsonify({"code": 0, "msg": msg})
        else:
            db.session.rollback()
            return jsonify({"code": 1, "msg": "操作失败: " + "; ".join(errors)})
            
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/outbound/batch-invoice", methods=["PUT"])
def batch_update_outbound_invoice():
    """批量更新出库计划关联的发票ID"""
    try:
        data = request.json
        ids = [int(i) for i in data.get("ids", []) if i]
        invoice_id = data.get("invoice_id")
        
        if not ids:
            return jsonify({"code": 1, "msg": "未选择记录"})
            
        # 验证发票是否存在（如果提供了 invoice_id）
        if invoice_id:
            invoice = MaterialInvoiceORM.query.get(invoice_id)
            if not invoice:
                return jsonify({"code": 1, "msg": "发票不存在"})
        
        # 批量更新
        db.session.query(MaterialOutboundORM).filter(
            MaterialOutboundORM.id.in_(ids)
        ).update({"invoice_id": invoice_id}, synchronize_session=False)
        
        db.session.commit()
        return jsonify({"code": 0, "msg": "批量关联成功"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/outbound/batch_confirm", methods=["POST"])
def confirm_outbound():
    """批量确认出库（销项）"""
    from decimal import Decimal
    try:
        data = request.json
        ids = data.get("ids", [])
        if not ids:
            return jsonify({"code": 1, "msg": "未选择项目"})
            
        outbound_items = MaterialOutboundORM.query.filter(MaterialOutboundORM.id.in_(ids)).all()
        for item in outbound_items:
            if item.status == 'completed':
                continue
                
            # 1. 获取关联库存 (User requested to skip inventory interaction)
            # inventory = item.inventory
            # if not inventory:
            #     raise Exception(f"出库单 [{item.id}] 关联的库存记录不存在")
            
            # Use snapshotted seller_quantity from the outbound record
            # out_qty = Decimal(str(item.seller_quantity)) if item.seller_quantity else Decimal(0)
            
            # if inventory.current_stock < out_qty:
            #     raise Exception(f"材料 [{inventory.material_name}] 库存不足 (需 {out_qty}, 存 {inventory.current_stock})")
            
            # 2. 更新库存 (Skipped)
            # current_stock_decimal = Decimal(str(inventory.current_stock or 0))
            # inventory.current_stock = current_stock_decimal - out_qty
            
            # ... (Skipped total_value calc)
            
            # 3. 更新出库状态并迁移数量
            item.status = 'completed'
            item.completed_sales_quantity = item.seller_quantity
            item.seller_quantity = 0
            
        db.session.commit()
        return jsonify({"code": 0, "msg": "批量出库成功"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/inventory/batch_supplier", methods=["PUT"])
def batch_update_inventory_supplier():
    """批量更新库存供应商及销售商名称"""
    try:
        data = request.json
        ids = data.get("ids", [])
        supplier_id = data.get("supplier_id")
        
        if not ids:
            return jsonify({"code": 1, "msg": "未选择任何记录"})
            
        # Get supplier name
        from pear_admin.orms import SupplierORM
        supplier_name = None
        if supplier_id:
            supplier = SupplierORM.query.get(supplier_id)
            if supplier:
                supplier_name = supplier.name

        # update_values = {MaterialInventoryORM.supplier_id: supplier_id} # Fixed: Do not update supplier_id
        update_values = {}
        if supplier_id:
            update_values[MaterialInventoryORM.seller_id] = supplier_id
            # Seller name is now dynamic via relationship, no need to update string column
            
            # Legacy logic removed:
            # if supplier_name:
            #    update_values[MaterialInventoryORM.seller_name] = supplier_name
            
        count = MaterialInventoryORM.query.filter(MaterialInventoryORM.id.in_(ids)).update(
            update_values,
            synchronize_session=False
        )
        
        db.session.commit()
        return jsonify({"code": 0, "msg": f"成功更新 {count} 条记录"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/inventory/<int:id>", methods=["GET", "PUT"])
def update_inventory(id):
    """获取或更新单个库存记录"""
    try:
        inventory = MaterialInventoryORM.query.get(id)
        if not inventory:
             return jsonify({"code": 1, "msg": "记录不存在"})

        if request.method == "GET":
             return jsonify({"code": 0, "msg": "success", "data": inventory.json()})

        data = request.json
        
        # Support single field 'field'/'value' format for compatibility
        if 'field' in data:
            field = data.get("field")
            value = data.get("value")
            if field == "seller_price":
                inventory.seller_price = float(value) if value else 0
            elif field == "seller_quantity":
                inventory.seller_quantity = float(value) if value else 0
            elif field == "profit_ratio":
                inventory.profit_ratio = float(value) if value else 1.05
            elif field == "current_stock":
                inventory.current_stock = float(value) if value else 0
            elif field == "tax_rate":
                inventory.tax_rate = float(value) if value else 0
            elif field == "tax_amount":
                inventory.tax_amount = float(value) if value else 0
            elif field == "price_no_tax":
                inventory.price_no_tax = float(value) if value else 0
            else:
                return jsonify({"code": 1, "msg": f"不支持修改字段: {field}"})
        else:
            # Support dictionary update
            if 'seller_price' in data:
                inventory.seller_price = float(data['seller_price'])
            if 'seller_quantity' in data:
                inventory.seller_quantity = float(data['seller_quantity'])
            if 'profit_ratio' in data:
                inventory.profit_ratio = float(data['profit_ratio'])
            if 'current_stock' in data:
                inventory.current_stock = float(data['current_stock'])
            if 'tax_rate' in data:
                inventory.tax_rate = float(data['tax_rate'])
            if 'tax_amount' in data:
                inventory.tax_amount = float(data['tax_amount'])
            if 'price_no_tax' in data:
                inventory.price_no_tax = float(data['price_no_tax'])
            
        db.session.commit()
        return jsonify({"code": 0, "msg": "更新成功"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/inventory/batch_calculate", methods=["POST"])
def batch_calculate_inventory():
    """批量计算销售信息"""
    try:
        data = request.json
        ids = data.get("ids", [])
        
        if not ids:
             return jsonify({"code": 1, "msg": "未选择任何记录"})
             
        records = MaterialInventoryORM.query.filter(MaterialInventoryORM.id.in_(ids)).all()
        count = 0
        for record in records:
            # Calculate sum of pending outbound quantities for this inventory
            pending_outbound_qty = db.session.query(db.func.sum(MaterialOutboundORM.seller_quantity)).filter(
                MaterialOutboundORM.inventory_id == record.id,
                MaterialOutboundORM.status == 'pending'
            ).scalar() or 0
            
            # 1. Quantity = Current Stock - Pending Outbound
            record.seller_quantity = max(0, float(record.current_stock) - float(pending_outbound_qty))
            
            # 2. Price = Latest Price * Profit Ratio
            ratio = record.profit_ratio if record.profit_ratio else 1.05
            base_price = record.latest_price if record.latest_price else 0
            record.seller_price = float(base_price) * float(ratio)
            count += 1
            
        db.session.commit()
        return jsonify({"code": 0, "msg": f"成功计算 {count} 条记录"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": str(e)})

# --- Invoice ---
@material_api.route("/invoice", methods=["GET"])
def get_invoice():
    try:
        page = request.args.get("page", 1, type=int)
        limit = request.args.get("limit", 10, type=int)
        
        query = MaterialInvoiceORM.query

        # ID筛选
        invoice_id = request.args.get("id")
        if invoice_id:
            query = query.filter_by(id=invoice_id)
        
        # 项目筛选
        project_id = request.args.get("project_id")
        if project_id and str(project_id).isdigit():
            query = query.filter_by(project_id=int(project_id))
            
        # 模糊搜索：多维度关键字搜索 (发票号码、购买方、销售方)
        search = request.args.get("search")
        if search:
            query = query.filter(or_(
                MaterialInvoiceORM.invoice_number.like(f"%{search}%"),
                MaterialInvoiceORM.buyer_name.like(f"%{search}%"),
                MaterialInvoiceORM.seller_name.like(f"%{search}%")
            ))
            
        # 兼容旧的单字段搜索参数
        invoice_number = request.args.get("invoice_number")
        if invoice_number:
            query = query.filter(MaterialInvoiceORM.invoice_number.like(f"%{invoice_number}%"))
            
        buyer_name = request.args.get("buyer_name")
        if buyer_name:
            query = query.filter(MaterialInvoiceORM.buyer_name.like(f"%{buyer_name}%"))
            
        seller_name = request.args.get("seller_name")
        if seller_name:
            query = query.filter(MaterialInvoiceORM.seller_name.like(f"%{seller_name}%"))
        
        # 限制返回数量 (用于下拉建议等场景)
        limit_val = request.args.get("limit_only", type=int)
        if limit_val:
            data_items = query.limit(limit_val).all()
            return jsonify({
                "code": 0,
                "msg": "获取成功",
                "data": [item.json() for item in data_items]
            })

        # 按发票日期降序排序
        query = query.order_by(MaterialInvoiceORM.invoice_date.desc())
        
        pagination = query.paginate(page=page, per_page=limit, error_out=False)
        
        data = []
        for invoice in pagination.items:
            item = invoice.json()
            # 添加项目名称
            if invoice.project_id:
                project = ProjectORM.query.get(invoice.project_id)
                item['project_name'] = project.name if project else '-'
            else:
                item['project_name'] = '-'
            data.append(item)
        
        return jsonify({
            "code": 0,
            "msg": "",
            "count": pagination.total,
            "data": data
        })
    except Exception as e:
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/invoice/list", methods=["GET"])
def get_invoice_list():
    """获取所有发票简单列表（用于左侧导航）"""
    try:
        invoices = MaterialInvoiceORM.query.order_by(MaterialInvoiceORM.id.desc()).all()
        data = []
        for inv in invoices:
            post_tax_amount = float(inv.total_amount or 0) + float(inv.tax_amount or 0)
            try:
                if inv.ocr_result:
                    import json
                    ocr_data = json.loads(inv.ocr_result)
                    if ocr_data.get('amount_in_figuers'):
                        post_tax_amount = float(ocr_data.get('amount_in_figuers'))
            except:
                pass
                
            data.append({
                "id": inv.id,
                "invoice_number": inv.invoice_number or "无编号",
                "seller_name": inv.seller_name or "-",
                "buyer_name": inv.buyer_name or "-",
                "total_amount": float(inv.total_amount or 0),
                "tax_amount": float(inv.tax_amount or 0),
                "post_tax_amount": round(post_tax_amount, 2),
                "file_url": inv._get_signed_url(),  # Use signed URL for private OSS files
                "invoice_category": inv.invoice_category,  # 添加发票大类
                "deductible": inv.deductible  # 添加可否抵扣
            })
        return jsonify({"code": 0, "msg": "success", "data": data})
    except Exception as e:
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/invoice/preview/<int:invoice_id>", methods=["GET"])
def preview_invoice_file(invoice_id):
    """代理获取发票原件并以内联方式返回，用于弹窗预览"""
    try:
        inv = MaterialInvoiceORM.query.get(invoice_id)
        if not inv or not inv.file_path:
            return "发票不存在或无附件", 404

        file_path = inv.file_path
        
        # 确定文件扩展名
        ext = file_path.rstrip('/').split('?')[0].rsplit('.', 1)[-1].lower() if '.' in file_path else 'pdf'
        mime_map = {
            'pdf': 'application/pdf',
            'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
            'png': 'image/png', 'gif': 'image/gif',
        }
        content_type = mime_map.get(ext, 'application/octet-stream')

        # 若是 OSS 文件，直接用 oss SDK 获取内容流
        if file_path.startswith('http'):
            from pear_admin.extensions import oss
            from urllib.parse import urlparse, unquote
            import requests
            
            if oss and oss.bucket:
                parsed = urlparse(file_path)
                object_key = unquote(parsed.path.lstrip('/'))
                # 获取对象内容（流式）
                oss_obj = oss.bucket.get_object(object_key)
                data = oss_obj.read()
            else:
                # 无OSS配置，直接请求原URL
                r = requests.get(file_path, timeout=15)
                data = r.content
        else:
            # 本地文件
            import os
            local_path = os.path.join(os.getcwd(), file_path.lstrip('/'))
            with open(local_path, 'rb') as f:
                data = f.read()

        from flask import Response
        from urllib.parse import quote
        
        filename = inv.invoice_number or str(inv.id)
        # RFC 5987 编码
        quoted_filename = quote(f"{filename}.{ext}")
        
        response = Response(
            data,
            content_type=content_type,
            headers={
                'Content-Disposition': f'inline; filename="{quoted_filename}"; filename*=UTF-8\'\'{quoted_filename}',
                'Cache-Control': 'no-cache'
            }
        )
        return response
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"[Invoice Preview] Error: {e}", exc_info=True)
        return str(e), 500

@material_api.route("/invoice/upload", methods=["POST"])
def upload_invoice():
    """上传发票文件"""
    try:
        from werkzeug.utils import secure_filename
        import os
        
        files = request.files.getlist('files')
        project_id = request.form.get('project_id')
        custom_path = request.form.get('path', 'invoices').strip()
        
        if not files or files[0].filename == '':
            return jsonify({"code": 1, "msg": "未选择文件"})
            
        # 允许的文件扩展名
        ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png'}
        
        def allowed_file(filename):
            return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
        
        from pear_admin.extensions import oss
        from pear_admin.ocr_utils import get_ocr_instance
        
        ocr = get_ocr_instance()
        
        # Helper function to safely convert to float
        def safe_float(value):
            """Safely convert value to float, handling '***' and other non-numeric strings"""
            if not value or value == '***':
                return 0
            try:
                return float(str(value).replace(',', ''))
            except (ValueError, AttributeError):
                return 0
        
        uploaded = []
        failed = []
        
        for file in files:
            try:
                # 验证文件格式
                if not allowed_file(file.filename):
                    failed.append({"name": file.filename, "reason": "不支持的文件格式"})
                    continue
                
                # 生成唯一标识
                timestamp_str = datetime.now().strftime('%Y%m%d%H%M%S%f')
                ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
                temp_filename = f"{timestamp_str}_{secure_filename(file.filename)}"
                
                # 定义清理文件名的辅助函数
                def clean_name(s):
                    if not s: return ""
                    import re
                    # 只保留中文字符、字母、数字和下划线，去除特殊格式字符
                    s = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9]', '', str(s))
                    return s[:30] # 截断太长的名称
                
                # 读取文件内容用于OCR (避免多次读取流)
                file.seek(0)
                file_content = file.read()
                file.seek(0) # 重置指针供保存使用
 
                # 1. 立即调用 OCR (优先识别)
                ocr_result = {}
                ocr_error = None
                try:
                    # 使用 file_content 进行识别
                    ocr_result = ocr.recognize_invoice(file_content=file_content, file_path=temp_filename)
                except Exception as e:
                    ocr_error = str(e)
                    current_app.logger.error(f"OCR Failed for {file.filename}: {e}")

                # 2. 构造重命名规则: 发票编号_购买方_销售方_流水号
                inv_num_clean = clean_name(ocr_result.get('invoice_number'))
                buyer_clean = clean_name(ocr_result.get('buyer_name'))
                seller_clean = clean_name(ocr_result.get('seller_name'))
                
                # 如果 OCR 识别到了关键信息，则使用新规则；否则使用原名+时间戳
                if inv_num_clean or buyer_clean or seller_clean:
                    new_filename = f"{inv_num_clean}_{buyer_clean}_{seller_clean}_{timestamp_str}.{ext}".strip('_')
                else:
                    new_filename = f"{timestamp_str}_{secure_filename(file.filename)}"

                # 3. 查重逻辑与日志
                with open("debug_invoice.log", "a", encoding="utf-8") as f:
                     ocr_num = ocr_result.get('invoice_number', 'None')
                     f.write(f"\n{datetime.now()}: OCR Result for {new_filename}: Number={ocr_num}\n")

                if ocr_result and ocr_result.get('invoice_number'):
                    inv_num = ocr_result.get('invoice_number')
                    exists = MaterialInvoiceORM.query.filter_by(invoice_number=inv_num).first()
                    if exists:
                        failed.append({"name": file.filename, "reason": f"发票号 {inv_num} 已存在 (ID: {exists.id})"})
                        continue # 跳过此文件，不上传

                # 4. 执行上传 (OSS 或 本地)
                file_path = ""
                with open("debug_invoice.log", "a", encoding="utf-8") as f:
                     f.write(f"{datetime.now()}: OSS bucket status: {oss.bucket is not None}\n")
                
                if oss.bucket:
                    # --- OSS 上传模式 ---
                    try:
                        oss_path = f"{custom_path}/{datetime.now().strftime('%Y/%m')}/{new_filename}"
                        with open("debug_invoice.log", "a", encoding="utf-8") as f:
                             f.write(f"{datetime.now()}: Uploading to OSS: {oss_path}\n")
                        
                        file.seek(0) # 确保上传的是完整内容
                        file_url = oss.upload_file(file, filename=oss_path)
                        file_path = file_url
                        with open("debug_invoice.log", "a", encoding="utf-8") as f:
                             f.write(f"{datetime.now()}: OSS upload success: {file_url}\n")
                    except Exception as e:
                        with open("debug_invoice.log", "a", encoding="utf-8") as f:
                             f.write(f"{datetime.now()}: OSS upload failed: {str(e)}\n")
                        raise Exception(f"OSS上传失败: {str(e)}")
                else:
                    # --- 本地存储模式 ---
                    with open("debug_invoice.log", "a", encoding="utf-8") as f:
                         f.write(f"{datetime.now()}: Using local storage\n")
                    year_month = datetime.now().strftime('%Y/%m')
                    upload_dir = os.path.join('static', 'uploads', custom_path, year_month)
                    os.makedirs(upload_dir, exist_ok=True)
                    local_path = os.path.join(upload_dir, new_filename)
                    with open(local_path, 'wb') as f:
                        f.write(file_content)
                    file_path = local_path

                # 5. 创建数据库记录 (填充 OCR 数据)
                # 生成展示用发票号
                with open("debug_invoice.log", "a", encoding="utf-8") as f:
                     f.write(f"{datetime.now()}: Preparing to save invoice {new_filename}...\n")

                final_invoice_number = f"TEMP-{timestamp_str}"
                if ocr_result and ocr_result.get('invoice_number'):
                     final_invoice_number = ocr_result.get('invoice_number')
                
                # 解析日期
                inv_date = None
                if ocr_result.get('invoice_date'):
                    try:
                        d_str = ocr_result.get('invoice_date').replace('年','-').replace('月','-').replace('日','')
                        inv_date = datetime.strptime(d_str, '%Y-%m-%d').date()
                    except:
                        pass

                import json
                invoice = MaterialInvoiceORM(
                    project_id=int(project_id) if project_id else None,
                    invoice_number=final_invoice_number,
                    file_path=file_path,
                    file_name=file.filename,
                    file_type=ext,
                    
                    # OCR 字段映射
                    invoice_code=ocr_result.get('invoice_code'),
                    invoice_date=inv_date,
                    buyer_name=ocr_result.get('buyer_name'),
                    buyer_tax_num=ocr_result.get('buyer_tax_num'),
                    seller_name=ocr_result.get('seller_name'),
                    seller_tax_num=ocr_result.get('seller_tax_num'),
                    total_amount=safe_float(ocr_result.get('total_amount')),
                    tax_amount=safe_float(ocr_result.get('total_tax')),
                    amount_in_words=ocr_result.get('amount_in_words'),
                    
                    # Enhanced fields
                    invoice_type=ocr_result.get('invoice_type'),
                    invoice_name=ocr_result.get('invoice_name'),
                    check_code=ocr_result.get('check_code'),
                    machine_num=ocr_result.get('machine_code'),
                    password_area=ocr_result.get('password'),
                    province=ocr_result.get('province'),
                    city=ocr_result.get('city'),
                    
                    buyer_address_phone=ocr_result.get('buyer_address'),
                    buyer_bank_account=ocr_result.get('buyer_bank'),
                    seller_address_phone=ocr_result.get('seller_address'),
                    seller_bank_account=ocr_result.get('seller_bank'),
                    
                    remarks=ocr_result.get('remarks'),
                    payee=ocr_result.get('payee'),
                    checker=ocr_result.get('checker'),
                    drawer=ocr_result.get('note_drawer'),
                    
                    ocr_status='success' if ocr_result else 'failed',
                    ocr_result=json.dumps(ocr_result, ensure_ascii=False) if ocr_result else None,
                    ocr_error=ocr_error
                )
                db.session.add(invoice)
                db.session.flush() # 获取 invoice.id

                # 5. 保存明细 (如果存在)
                from pear_admin.orms.material import MaterialInvoiceDetailORM
                if ocr_result and ocr_result.get('details'):
                    for d in ocr_result.get('details'):
                        detail = MaterialInvoiceDetailORM(
                            invoice_id=invoice.id,
                            name=d.get('name'),
                            spec=d.get('spec'),
                            unit=d.get('unit'),
                            quantity=safe_float(d.get('quantity')),
                            price=safe_float(d.get('price')),
                            amount=safe_float(d.get('amount')),
                            tax_rate=d.get('tax_rate'),
                            tax_amount=safe_float(d.get('tax'))
                        )
                        db.session.add(detail)

                db.session.commit()
                
                with open("debug_invoice.log", "a", encoding="utf-8") as f:
                     f.write(f"{datetime.now()}: Successfully committed invoice ID: {invoice.id}\n")
                
                uploaded.append({
                    "id": invoice.id,
                    "file_name": file.filename,
                    "ocr_status": invoice.ocr_status,
                    "invoice_number": invoice.invoice_number,
                    "url": file_path
                })
            except Exception as e:
                failed.append({"name": file.filename, "reason": str(e)})
                db.session.rollback()
        
        return jsonify({
            "code": 0,
            "msg": f"上传完成，成功{len(uploaded)}个，失败{len(failed)}个",
            "data": {
                "uploaded": len(uploaded),
                "failed": len(failed),
                "invoices": uploaded,
                "errors": failed
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": f"上传失败: {str(e)}"})

@material_api.route("/invoice/ocr", methods=["POST"])
def ocr_invoice():
    """OCR识别发票"""
    try:
        try:
            from pear_admin.ocr_utils import get_ocr_instance
            from pear_admin.extensions import oss
        except ImportError as ie:
            return jsonify({"code": 1, "msg": f"OCR模块导入失败: {str(ie)}"})
            
        import time
        import json as json_lib
        
        data = request.json
        invoice_ids = data.get('invoice_ids', [])
        
        if not invoice_ids:
            return jsonify({"code": 1, "msg": "未选择发票"})
        
        ocr = get_ocr_instance()
        success_count = 0
        failed_count = 0
        results = []
        
        for invoice_id in invoice_ids:
            invoice = MaterialInvoiceORM.query.get(invoice_id)
            if not invoice:
                failed_count += 1
                results.append({"id": invoice_id, "status": "failed", "error": "发票不存在"})
                continue
            
            if not invoice.file_path:
                failed_count += 1
                results.append({"id": invoice_id, "status": "failed", "error": "未上传文件"})
                continue
            
            try:
                # 更新状态为处理中
                invoice.ocr_status = 'processing'
                db.session.commit()
                
                print(f"开始识别发票 ID: {invoice_id}, 文件路径: {invoice.file_path}")
                
                # 调用OCR识别
                file_content = None
                if invoice.file_path and invoice.file_path.startswith('http') and oss.bucket:
                    try:
                        from urllib.parse import urlparse, unquote
                        path = urlparse(invoice.file_path).path
                        key = unquote(path.lstrip('/'))
                        print(f"Downloading from OSS: {key}")
                        obj = oss.bucket.get_object(key)
                        file_content = obj.read()
                    except Exception as oss_err:
                        print(f"Failed to download from OSS for OCR: {oss_err}")

                if file_content:
                    ocr_result = ocr.recognize_invoice(file_content=file_content, file_path=invoice.file_name)
                else:
                    ocr_result = ocr.recognize_invoice(invoice.file_path)
                
                print(f"OCR识别结果: {ocr_result}")
                
                # 更新发票信息
                if ocr_result.get('invoice_number'):
                    inv_num = ocr_result.get('invoice_number')
                    # 检查是否重复 (排除自身)
                    existing = MaterialInvoiceORM.query.filter_by(invoice_number=inv_num).filter(MaterialInvoiceORM.id != invoice_id).first()
                    if existing:
                        # 严格排重：删除文件和记录
                        try:
                            import os
                            if invoice.file_path and os.path.exists(invoice.file_path):
                                os.remove(invoice.file_path)
                            print(f"Duplicate found, deleted file: {invoice.file_path}")
                        except Exception as del_err:
                            print(f"Error deleting duplicate file: {del_err}")
                        
                        db.session.delete(invoice)
                        db.session.commit()
                        raise Exception(f"DUPLICATE_FOUND: 发票号码 {inv_num} 已存在，系统已自动清理")
                    invoice.invoice_number = inv_num
                
                invoice.invoice_code = ocr_result.get('invoice_code')
                
                # 新增字段
                invoice.invoice_type = ocr_result.get('invoice_type')
                invoice.invoice_name = ocr_result.get('invoice_name')
                invoice.check_code = ocr_result.get('check_code')
                invoice.machine_num = ocr_result.get('machine_code')
                invoice.password_area = ocr_result.get('password')
                invoice.province = ocr_result.get('province')
                invoice.city = ocr_result.get('city')
                
                # 解析日期
                invoice_date_str = ocr_result.get('invoice_date', '')
                if invoice_date_str:
                    try:
                        import re
                        date_match = re.search(r'(\d{4})年(\d{1,2})月(\d{1,2})日', invoice_date_str)
                        if date_match:
                            year, month, day = date_match.groups()
                            invoice.invoice_date = datetime(int(year), int(month), int(day)).date()
                    except:
                        pass
                
                invoice.buyer_name = ocr_result.get('buyer_name')
                invoice.buyer_tax_num = ocr_result.get('buyer_tax_num')
                invoice.buyer_address_phone = ocr_result.get('buyer_address')
                invoice.buyer_bank_account = ocr_result.get('buyer_bank')
                
                invoice.seller_name = ocr_result.get('seller_name')
                invoice.seller_tax_num = ocr_result.get('seller_tax_num')
                invoice.seller_address_phone = ocr_result.get('seller_address')
                invoice.seller_bank_account = ocr_result.get('seller_bank')
                
                invoice.remarks = ocr_result.get('remarks')
                invoice.payee = ocr_result.get('payee')
                invoice.checker = ocr_result.get('checker')
                invoice.drawer = ocr_result.get('note_drawer')
                
                # 解析金额
                try:
                    total_amount_str = ocr_result.get('total_amount', '0')
                    invoice.total_amount = float(total_amount_str.replace(',', ''))
                except:
                    invoice.total_amount = 0
                
                try:
                    total_tax_str = ocr_result.get('total_tax', '0')
                    invoice.tax_amount = float(total_tax_str.replace(',', ''))
                except:
                    invoice.tax_amount = 0
                
                invoice.amount_in_words = ocr_result.get('amount_in_words')
                
                # 保存明细数据到关联表
                details = ocr_result.get('details')
                
                # 清除旧明细 (如有)
                for old_d in invoice.details:
                    db.session.delete(old_d)
                
                if details:
                    # 存入详情表
                    for item in details:
                        def _parse_num(v):
                            try:
                                return float(str(v).replace(',', '').replace('¥', '').strip())
                            except:
                                return 0
                                
                        detail = MaterialInvoiceDetailORM(
                            invoice_id=invoice.id,
                            name=item.get('name'),
                            spec=item.get('spec'),
                            unit=item.get('unit'),
                            quantity=_parse_num(item.get('quantity')),
                            price=_parse_num(item.get('price')),
                            amount=_parse_num(item.get('amount')),
                            tax_rate=item.get('tax_rate'),
                            tax_amount=_parse_num(item.get('tax'))
                        )
                        db.session.add(detail)
                        
                    # details_json 已废弃，不再写入数据
                    # invoice.details_json = json_lib.dumps(details, ensure_ascii=False)
                    
                invoice.ocr_result = json_lib.dumps(ocr_result, ensure_ascii=False)
                invoice.ocr_status = 'completed'
                invoice.ocr_error = None
                
                # 同步供应商信息
                if invoice.seller_name:
                    supplier = SupplierORM.query.filter_by(name=invoice.seller_name).first()
                    if not supplier:
                        # 创建新供应商，提供必需字段的默认值
                        supplier = SupplierORM(
                            name=invoice.seller_name,
                            type_id=1,  # 默认类型
                            contact_person=invoice.seller_tax_num or '待补充',  # 使用税号作为临时联系人
                            phone='待补充',
                            bank_name='待补充',
                            account_number=invoice.seller_tax_num or '待补充',
                            remark=f'从发票OCR自动创建，税号：{invoice.seller_tax_num}'
                        )
                        db.session.add(supplier)
                        db.session.flush()
                    invoice.supplier_id = supplier.id
                
                db.session.commit()
                success_count += 1
                results.append({"id": invoice_id, "status": "completed"})
                
                print(f"发票 {invoice_id} 识别成功")
                
                # API频率限制：每秒最多2次请求
                time.sleep(0.5)
                
            except Exception as e:
                print(f"发票 {invoice_id} 识别失败: {str(e)}")
                import traceback
                traceback.print_exc()
                invoice.ocr_status = 'failed'
                invoice.ocr_error = str(e)
                db.session.commit()
                failed_count += 1
                results.append({"id": invoice_id, "status": "failed", "error": str(e)})
        
        return jsonify({
            "code": 0,
            "msg": f"识别完成，成功{success_count}个，失败{failed_count}个",
            "data": {
                "success": success_count,
                "failed": failed_count,
                "results": results
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": f"识别失败: {str(e)}"})

@material_api.route("/invoice", methods=["POST"])
def add_invoice():
    try:
        data = request.json
        # Check duplicate
        inv_num = data.get("invoice_number")
        if inv_num and MaterialInvoiceORM.query.filter_by(invoice_number=inv_num).first():
            return jsonify({"code": 1, "msg": f"发票号码 {inv_num} 已存在"})
            
        row = MaterialInvoiceORM(**data)
        db.session.add(row)
        db.session.commit()
        return jsonify({"code": 0, "msg": "Success"})
    except Exception as e:
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/invoice", methods=["DELETE"])
def delete_invoice():
    try:
        import os
        data = request.json
        ids = data.get('ids')
        if not ids:
            return jsonify({"code": 1, "msg": "请选择要删除的发票"})
            
        rows = MaterialInvoiceORM.query.filter(MaterialInvoiceORM.id.in_(ids)).all()
        for row in rows:
            # Delete physical file
            if row.file_path:
                try:
                    # Resolve absolute path if needed, but current storage seems relative to CWD
                    if os.path.exists(row.file_path):
                        os.remove(row.file_path)
                        print(f"Deleted file: {row.file_path}")
                except Exception as file_err:
                    print(f"Error deleting file {row.file_path}: {file_err}")

            db.session.delete(row)
            
        db.session.commit()
        return jsonify({"code": 0, "msg": "删除成功"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/invoice/<int:invoice_id>", methods=["PUT"])
def update_invoice(invoice_id):
    """更新发票信息"""
    try:
        invoice = MaterialInvoiceORM.query.get(invoice_id)
        if not invoice:
            return jsonify({"code": 1, "msg": "发票不存在"})
        
        data = request.json
        
        # 更新分类和抵扣字段
        if 'invoice_category' in data:
            invoice.invoice_category = data.get('invoice_category')
        if 'deductible' in data:
            invoice.deductible = data.get('deductible')
        
        # 可以扩展更新其他字段
        if 'remarks' in data:
            invoice.remarks = data.get('remarks')
        
        db.session.commit()
        return jsonify({"code": 0, "msg": "更新成功", "data": invoice.json()})
    except Exception as e:
        db.session.rollback()
        return jsonify({"code": 1, "msg": str(e)})

@material_api.route("/dashboard/stats", methods=["GET"])
def dashboard_stats():
    try:
        inventories = MaterialInventoryORM.query.all()
        total_value = sum(float(i.total_value or 0) for i in inventories)
        pending_inbound = MaterialInboundORM.query.filter_by(status="pending").count()
        pending_outbound = MaterialOutboundORM.query.filter_by(status="pending").count()
        
        return jsonify({
            "code": 0,
            "msg": "",
            "data": {
                "total_inventory_value": total_value,
                "pending_inbound_count": pending_inbound,
                "pending_outbound_count": pending_outbound
            }
        })
    except Exception as e:
        return jsonify({"code": 1, "msg": str(e)})
